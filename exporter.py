# exporter.py

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_shopee_excel(products: list, filepath: str) -> bool:
    """
    Exports a list of product dictionaries to a Shopee Mass Upload compatible Excel file.
    Each product dict should contain details like title, description, price, stock, weight, image_url, etc.
    """
    # Shopee Mass Upload standard headers (4-row format)
    # Row 1-2: System description & Instructions
    # Row 3: Technical column names
    # Row 4: Human readable labels
    
    headers_technical = [
        "category_id", "product_name", "product_description", "parent_sku", 
        "price", "stock", "product_weight", "days_to_ship",
        "image_1", "image_2", "image_3", "image_4", "image_5", "image_6", "image_7", "image_8", "image_9"
    ]
    
    headers_human = [
        "Category ID", "Product Name", "Product Description", "Parent SKU",
        "Price", "Stock", "Weight (kg)", "Days to Ship (Pre-order)",
        "Image 1", "Image 2", "Image 3", "Image 4", "Image 5", "Image 6", "Image 7", "Image 8", "Image 9"
    ]
    
    # Construct rows
    rows = []
    
    # System identification rows (standard for Shopee templates)
    rows.append(["Shopee Mass Upload Template v2.0"] + [""] * (len(headers_technical) - 1))
    rows.append(["Please do not modify the structure of this sheet. Fill in product details below."] + [""] * (len(headers_technical) - 1))
    rows.append(headers_technical)
    rows.append(headers_human)
    
    # Fill product data
    for idx, prod in enumerate(products):
        # Extract images
        sub_imgs = prod.get("sub_images", [])
        img_urls = [""] * 9
        for i in range(min(len(sub_imgs), 9)):
            img_urls[i] = sub_imgs[i]
            
        # If no sub_images but has main image
        if not img_urls[0] and prod.get("image_url"):
            img_urls[0] = prod.get("image_url")
            
        # Unique Parent SKU generation
        parent_sku = prod.get("parent_sku")
        if not parent_sku:
            source_prefix = "OVY" if prod.get("source") == "Olive Young" else "DSO"
            parent_sku = f"{source_prefix}_{prod.get('goods_no')}"
            
        row = [
            prod.get("category_id", 100000),      # Default Category (e.g. 100000 Beauty & Personal Care)
            prod.get("translated_title", prod.get("name")),
            prod.get("translated_description", prod.get("description_text", "")),
            parent_sku,
            prod.get("calculated_price", prod.get("sale_price", 0)),
            prod.get("stock", 100),                # Default stock
            prod.get("weight", 0.1),               # Weight in kg
            prod.get("days_to_ship", ""),          # Default non-preorder (empty)
            img_urls[0], img_urls[1], img_urls[2], img_urls[3], img_urls[4], img_urls[5], img_urls[6], img_urls[7], img_urls[8]
        ]
        rows.append(row)
        
    try:
        # Create a new workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Write rows
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply styling for headers
                if row_idx in [1, 2]:
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="EE4D2D", end_color="EE4D2D", fill_type="solid") # Shopee orange
                    if col_idx == 1:
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers_technical))
                elif row_idx in [3, 4]:
                    cell.font = Font(name="Arial", size=10, bold=True, color="333333")
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = Font(name="Arial", size=10)
                    
        # Apply borders and auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Error saving Shopee Excel file: {e}")
        return False

def export_to_shopify_csv(products: list, filepath: str) -> bool:
    """
    Exports a list of product dictionaries to a Shopify compatible CSV file.
    """
    columns = [
        "Handle", "Title", "Body (HTML)", "Vendor", "Product Category", "Type", "Tags", "Published",
        "Option1 Name", "Option1 Value", "Option2 Name", "Option2 Value", "Option3 Name", "Option3 Value",
        "Variant SKU", "Variant Grams", "Variant Inventory Tracker", "Variant Inventory Qty",
        "Variant Inventory Policy", "Variant Fulfillment Service", "Variant Price", "Variant Compare At Price",
        "Variant Requires Shipping", "Variant Taxable", "Variant Barcode", "Image Src", "Image Position",
        "Image Alt Text", "Gift Card", "SEO Title", "SEO Description", "Google Shopping / Google Product Category",
        "Google Shopping / Gender", "Google Shopping / Age Group", "Google Shopping / MPN",
        "Google Shopping / Condition", "Google Shopping / Custom Product", "Google Shopping / Custom Label 0",
        "Google Shopping / Custom Label 1", "Google Shopping / Custom Label 2", "Google Shopping / Custom Label 3",
        "Google Shopping / Custom Label 4", "Variant Image", "Variant Weight Unit", "Variant Tax Code", "Cost per item"
    ]
    
    rows = []
    for prod in products:
        # Create handle from title
        title = prod.get("translated_title", prod.get("name"))
        handle = title.lower().replace(" ", "-").replace("/", "-").replace("&", "and")
        handle = ''.join(c for c in handle if c.isalnum() or c == '-')
        
        # Tags formatting
        tags = ", ".join(prod.get("tags", []))
        if not tags:
            tags = "K-Beauty, Cosmetics"
            
        # Image Src (Shopify uses main image)
        img_src = prod.get("image_url", "")
        
        # Price and original price
        price = prod.get("calculated_price", prod.get("sale_price", 0))
        compare_at = prod.get("original_price", price)
        if compare_at <= price:
            compare_at = "" # Don't show compare price if it's not higher
            
        # SKU
        source_prefix = "OVY" if prod.get("source") == "Olive Young" else "DSO"
        sku = f"{source_prefix}_{prod.get('goods_no')}"
        
        row = {
            "Handle": handle,
            "Title": title,
            "Body (HTML)": prod.get("description_html", prod.get("translated_description", "")),
            "Vendor": prod.get("brand_english", prod.get("brand")),
            "Product Category": "Health & Beauty > Personal Care > Cosmetics",
            "Type": "Cosmetics",
            "Tags": tags,
            "Published": "true",
            "Option1 Name": "Title",
            "Option1 Value": "Default Title",
            "Option2 Name": "",
            "Option2 Value": "",
            "Option3 Name": "",
            "Option3 Value": "",
            "Variant SKU": sku,
            "Variant Grams": round(prod.get("weight", 0.1) * 1000), # Shopify uses grams
            "Variant Inventory Tracker": "shopify",
            "Variant Inventory Qty": prod.get("stock", 100),
            "Variant Inventory Policy": "deny",
            "Variant Fulfillment Service": "manual",
            "Variant Price": price,
            "Variant Compare At Price": compare_at,
            "Variant Requires Shipping": "true",
            "Variant Taxable": "true",
            "Variant Barcode": "",
            "Image Src": img_src,
            "Image Position": 1,
            "Image Alt Text": title,
            "Gift Card": "false",
            "SEO Title": title,
            "SEO Description": prod.get("translated_description", "")[:160],
            "Google Shopping / Google Product Category": "",
            "Google Shopping / Gender": "Unisex",
            "Google Shopping / Age Group": "Adult",
            "Google Shopping / MPN": "",
            "Google Shopping / Condition": "New",
            "Google Shopping / Custom Product": "false",
            "Google Shopping / Custom Label 0": "",
            "Google Shopping / Custom Label 1": "",
            "Google Shopping / Custom Label 2": "",
            "Google Shopping / Custom Label 3": "",
            "Google Shopping / Custom Label 4": "",
            "Variant Image": "",
            "Variant Weight Unit": "g",
            "Variant Tax Code": "",
            "Cost per item": prod.get("sale_price", 0) # Korean cost
        }
        rows.append(row)
        
    try:
        df = pd.DataFrame(rows, columns=columns)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"Error exporting Shopify CSV: {e}")
        return False
